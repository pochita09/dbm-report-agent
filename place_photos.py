"""
Step 2: 写真配置ロジック

設計原則:
- 列幅・行高さ・デフォルト値・フォント情報はExcelのXMLから直接取得する
- openpyxlの column_dimensions / row_dimensions / _named_styles には依存しない
- openpyxlはワークブックの読み書き（load_workbook/save）と画像配置（OneCellAnchor）のみに使用
- OneCellAnchorで位置+画像EMUサイズを指定して配置（ストレッチ防止）
- アスペクト比を維持してセル内に収まるよう縮小、中央配置（トリミングなし）
"""

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage
import statistics
import struct
import zlib
import os
import shutil
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


# ============================================================
# XMLベースのテンプレート情報取得
# ============================================================

# Excel OpenXML名前空間
_NS = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

# 標準フォント(11pt)ごとのMaximum Digit Width
_MDW_TABLE = {
    'ＭＳ Ｐゴシック': 8, 'MS PGothic': 8,
    'ＭＳ ゴシック': 7, 'MS Gothic': 7,
    '游ゴシック': 8, 'Yu Gothic': 8, 'Yu Gothic UI': 7,
    'メイリオ': 9, 'Meiryo': 9, 'Meiryo UI': 8,
    'Arial': 7, 'Calibri': 7, 'Aptos': 7,
    'Times New Roman': 7, 'Verdana': 8, 'Tahoma': 7,
}


def parse_workbook_xml(template_path):
    """ワークブック全体のXML情報を取得する

    Returns:
        dict: {
            "sheet_paths": {"シート名": "xl/worksheets/sheet1.xml", ...},
            "mdw": int (Maximum Digit Width),
            "font_name": str (標準フォント名),
        }
    """
    with zipfile.ZipFile(template_path) as z:
        # workbook.xml.rels からシート名→XMLパスのマッピングを取得
        with z.open('xl/_rels/workbook.xml.rels') as f:
            rels_tree = ET.parse(f)
        rels_root = rels_tree.getroot()
        # 名前空間はデフォルト名前空間として定義されている場合がある
        rels_ns_uri = rels_root.tag.split('}')[0] + '}' if '}' in rels_root.tag else ''
        rid_to_path = {}
        for rel in rels_root.findall(f'{rels_ns_uri}Relationship'):
            rid = rel.get('Id')
            target = rel.get('Target')
            if target and 'worksheets/' in target:
                # 相対パスを正規化
                if not target.startswith('xl/'):
                    target = 'xl/' + target
                rid_to_path[rid] = target

        # workbook.xml からシート名→rIdのマッピングを取得
        with z.open('xl/workbook.xml') as f:
            wb_tree = ET.parse(f)
        wb_root = wb_tree.getroot()
        # 名前空間を動的に取得
        wb_ns_uri = wb_root.tag.split('}')[0] + '}' if '}' in wb_root.tag else ''
        r_ns_uri = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        sheet_paths = {}
        for sheet in wb_root.iter(f'{wb_ns_uri}sheet'):
            name = sheet.get('name')
            rid = sheet.get(f'{{{r_ns_uri}}}id')
            if rid and rid in rid_to_path:
                sheet_paths[name] = rid_to_path[rid]

        # styles.xml から標準フォント名を取得してMDWを決定
        mdw = 7  # デフォルト
        font_name = ''
        try:
            with z.open('xl/styles.xml') as f:
                styles_tree = ET.parse(f)
            styles_ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            # fonts要素の最初のfontが標準フォント
            fonts = styles_tree.findall('.//s:fonts/s:font', styles_ns)
            if fonts:
                name_elem = fonts[0].find('s:name', styles_ns)
                if name_elem is not None:
                    font_name = name_elem.get('val', '')
                    if font_name in _MDW_TABLE:
                        mdw = _MDW_TABLE[font_name]
        except (KeyError, ET.ParseError):
            pass

    return {
        "sheet_paths": sheet_paths,
        "mdw": mdw,
        "font_name": font_name,
    }


def parse_sheet_xml(template_path, sheet_xml_path):
    """シートのXMLから列幅・行高さ・デフォルト値・結合セル情報を直接取得する

    Args:
        template_path: XLSXファイルパス
        sheet_xml_path: ZIP内のシートXMLパス（例: "xl/worksheets/sheet3.xml"）

    Returns:
        dict: {
            "col_widths": {col_number: width_chars, ...},  # min-max展開済み
            "row_heights": {row_number: height_pt, ...},   # 全行（XML記載分）
            "default_col_width": float,
            "default_row_height": float,
            "merged_ranges": [(min_row, min_col, max_row, max_col), ...],
        }
    """
    with zipfile.ZipFile(template_path) as z:
        with z.open(sheet_xml_path) as f:
            tree = ET.parse(f)
    root = tree.getroot()

    # sheetFormatPr からデフォルト値を取得
    default_col_width = 8.43
    default_row_height = 15.0
    fmt = root.find('.//s:sheetFormatPr', _NS)
    if fmt is not None:
        dcw = fmt.get('defaultColWidth')
        if dcw:
            default_col_width = float(dcw)
        drh = fmt.get('defaultRowHeight')
        if drh:
            default_row_height = float(drh)

    # col要素から列幅を取得（min-maxを展開）
    col_widths = {}
    for col_elem in root.findall('.//s:col', _NS):
        col_min = int(col_elem.get('min', '0'))
        col_max = int(col_elem.get('max', '0'))
        width = col_elem.get('width')
        if width and col_min > 0:
            w = float(width)
            for c in range(col_min, col_max + 1):
                col_widths[c] = w

    # row要素から行高さを取得
    row_heights = {}
    for row_elem in root.findall('.//s:row', _NS):
        r = int(row_elem.get('r', '0'))
        ht = row_elem.get('ht')
        if ht and r > 0:
            row_heights[r] = float(ht)

    # mergeCells から結合セル情報を取得
    merged_ranges = []
    for mc in root.findall('.//s:mergeCell', _NS):
        ref = mc.get('ref', '')
        if ':' in ref:
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(ref)
            merged_ranges.append((min_row, min_col, max_row, max_col))

    return {
        "col_widths": col_widths,
        "row_heights": row_heights,
        "default_col_width": default_col_width,
        "default_row_height": default_row_height,
        "merged_ranges": merged_ranges,
    }


# ============================================================
# EMU変換ユーティリティ
# ============================================================

def col_width_to_emu(width_chars, mdw):
    """Excel公式の列幅→ピクセル→EMU変換"""
    if width_chars == 0:
        return 0
    px = int(((256 * width_chars + int(128 / mdw)) / 256) * mdw)
    return px * 9525


def row_height_to_emu(height_pt):
    return int(height_pt * 12700)


# ============================================================
# テンプレート解析
# ============================================================

def get_merged_cell_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return str(cell.value).strip()
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            if top_left.value is not None:
                return str(top_left.value).strip()
    return ""


def detect_photo_slots(ws, sheet_info=None):
    """写真スロットを検出する

    Args:
        ws: openpyxlのワークシートオブジェクト（セル値取得に使用）
        sheet_info: parse_sheet_xmlの返り値（None時は後方互換のためopenpyxlにフォールバック）

    Returns:
        tuple: (slots, row_heights, col_widths)
    """
    max_row = ws.max_row
    max_col = ws.max_column

    if sheet_info:
        # XMLベース: デフォルト値で埋めた完全な辞書を構築
        default_rh = sheet_info["default_row_height"]
        default_cw = sheet_info["default_col_width"]
        xml_row_heights = sheet_info["row_heights"]
        xml_col_widths = sheet_info["col_widths"]

        row_heights = {r: xml_row_heights.get(r, default_rh) for r in range(1, max_row + 1)}
        col_widths = {c: xml_col_widths.get(c, default_cw) for c in range(1, max_col + 1)}
    else:
        # フォールバック（後方互換）
        row_heights = {}
        col_widths = {}
        for r in range(1, max_row + 1):
            rd = ws.row_dimensions.get(r)
            row_heights[r] = rd.height if (rd and rd.height is not None) else 15.0
        for c in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            cd = ws.column_dimensions.get(col_letter)
            col_widths[c] = cd.width if (cd and cd.width is not None) else 8.43

    heights = list(row_heights.values())
    median_h = statistics.median(heights)
    threshold_h = median_h * 3.0
    photo_rows = sorted([r for r, h in row_heights.items() if h >= threshold_h])

    if not photo_rows:
        return [], row_heights, col_widths

    widths = list(col_widths.values())
    median_w = statistics.median(widths)
    # 中央値の半分以下の列を「区切り列」として除外し、残りをコンテンツ列とする
    # これにより列幅が均一なテンプレートでも正しく検出できる
    threshold_w = median_w * 0.5
    content_cols = sorted([c for c, w in col_widths.items() if w > threshold_w])

    if not content_cols:
        return [], row_heights, col_widths

    slots = []
    for idx, r in enumerate(photo_rows):
        row_min = photo_rows[idx - 1] + 1 if idx > 0 else 1
        for c in content_cols:
            # カテゴリ探索（写真行から最大3行上まで）
            category = ""
            cat_row = r
            for look_up in range(1, 4):
                search_row = r - look_up
                if search_row < row_min:
                    break
                candidate = get_merged_cell_value(ws, search_row, c)
                if candidate:
                    category = candidate
                    cat_row = search_row
                    break

            # カテゴリが見つからなければスロットから除外
            if not category:
                continue

            # セクション探索（カテゴリ行から最大3行上まで）
            section = ""
            for look_up in range(1, 4):
                search_row = cat_row - look_up
                if search_row < row_min:
                    break
                candidate = get_merged_cell_value(ws, search_row, c)
                if not candidate:
                    candidate = get_merged_cell_value(ws, search_row, content_cols[0])
                if candidate and candidate != category:
                    section = candidate
                    break

            slots.append({
                "row": r,
                "col": c,
                "category": category,
                "section": section,
                "row_heights": row_heights,
                "col_widths": col_widths,
            })

    return slots, row_heights, col_widths


# ============================================================
# スロット結合範囲取得
# ============================================================

def get_slot_merged_range(ws, row, col, merged_ranges=None):
    """スロットの結合範囲を取得する

    Args:
        ws: openpyxlのワークシートオブジェクト
        row, col: スロットの位置
        merged_ranges: parse_sheet_xmlのmerged_ranges（None時はopenpyxlにフォールバック）
    """
    if merged_ranges:
        for min_row, min_col, max_row, max_col in merged_ranges:
            if min_row <= row <= max_row and min_col <= col <= max_col:
                return (min_row, min_col, max_row, max_col)
        return (row, col, row, col)

    # フォールバック
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            return (merged_range.min_row, merged_range.min_col,
                    merged_range.max_row, merged_range.max_col)
    return (row, col, row, col)


# ============================================================
# ダミーPNG生成
# ============================================================

def make_dummy_png(filepath, width=400, height=300, color=(200, 100, 50), label=""):
    def png_chunk(name, data):
        c = struct.pack('>I', len(data)) + name + data
        crc = zlib.crc32(name + data) & 0xffffffff
        return c + struct.pack('>I', crc)

    header = b'\x89PNG\r\n\x1a\n'
    ihdr_data = struct.pack('>IIBBBBB', width, height, 8, 2, 0, 0, 0)
    ihdr = png_chunk(b'IHDR', ihdr_data)

    raw_rows = []
    for y in range(height):
        row = b'\x00'
        row += bytes(color) * width
        raw_rows.append(row)

    raw_data = b''.join(raw_rows)
    compressed = zlib.compress(raw_data)
    idat = png_chunk(b'IDAT', compressed)
    iend = png_chunk(b'IEND', b'')

    with open(filepath, 'wb') as f:
        f.write(header + ihdr + idat + iend)

    print(f"  ダミー画像生成: {filepath}")


# ============================================================
# 写真配置メイン
# ============================================================

def place_photos(template_path, output_path, photo_paths):
    import tempfile

    print(f"\n{'='*60}")
    print(f"写真配置開始: {Path(template_path).name}")
    print(f"{'='*60}")

    # XMLからワークブック情報を取得（openpyxlの内部APIに依存しない）
    wb_info = parse_workbook_xml(template_path)
    mdw = wb_info["mdw"]
    print(f"  MDW={mdw}（XMLから取得: フォント '{wb_info['font_name']}'）")

    wb = openpyxl.load_workbook(template_path)
    tmp_dir = tempfile.mkdtemp()
    total_placed = 0

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # XMLからシート情報を取得
            sheet_xml_path = wb_info["sheet_paths"].get(sheet_name)
            sheet_info = None
            if sheet_xml_path:
                sheet_info = parse_sheet_xml(template_path, sheet_xml_path)

            slots, row_heights, col_widths = detect_photo_slots(ws, sheet_info)

            if not slots:
                print(f"  スキップ: '{sheet_name}'")
                continue

            # 既存画像を全て削除（重複配置を防止）
            existing_count = len(ws._images)
            if existing_count > 0:
                ws._images.clear()
                print(f"  シート '{sheet_name}': 既存画像{existing_count}枚を削除")

            print(f"\n  シート '{sheet_name}': {len(slots)}スロットに配置")

            # 結合セル情報（XMLベース）
            merged_ranges = sheet_info["merged_ranges"] if sheet_info else None

            photo_iter = iter(photo_paths)

            for i, slot in enumerate(slots):
                try:
                    img_path = next(photo_iter)
                except StopIteration:
                    print(f"  [!] 写真が足りません（スロット{i+1}以降はスキップ）")
                    break

                # 未割り当てスロット（None）はスキップ
                if img_path is None:
                    print(f"    スロット{i+1}: (未割り当て)")
                    continue

                row_heights = slot["row_heights"]
                col_widths = slot["col_widths"]

                min_row, min_col, max_row, max_col = get_slot_merged_range(
                    ws, slot["row"], slot["col"], merged_ranges
                )

                slot_w_emu = sum(col_width_to_emu(col_widths.get(c, 8.43), mdw) for c in range(min_col, max_col + 1))
                slot_h_emu = sum(row_height_to_emu(row_heights.get(r, 15.0)) for r in range(min_row, max_row + 1))

                # --- アスペクト比維持フィット（トリミングなし・中央配置） ---
                with PILImage.open(img_path) as img:
                    img_w, img_h = img.size

                img_ratio = img_w / img_h
                slot_ratio = slot_w_emu / slot_h_emu

                if img_ratio > slot_ratio:
                    # 画像のほうが横長 → 幅に合わせる
                    fit_w_emu = slot_w_emu
                    fit_h_emu = int(slot_w_emu / img_ratio)
                else:
                    # 画像のほうが縦長 → 高さに合わせる
                    fit_h_emu = slot_h_emu
                    fit_w_emu = int(slot_h_emu * img_ratio)

                # EMU→px変換は切り捨てで誤差が出るため、アスペクト比をpx段階で再計算
                fit_w_px = max(1, fit_w_emu // 9525)
                fit_h_px = max(1, round(fit_w_px * img_h / img_w) if img_ratio > slot_ratio else fit_h_emu // 9525)
                # EMUをpxから逆算して一致させる
                fit_w_emu = fit_w_px * 9525
                fit_h_emu = fit_h_px * 9525

                # オフセットはEMU確定後に計算（中央配置）
                x_off = (slot_w_emu - fit_w_emu) // 2
                y_off = (slot_h_emu - fit_h_emu) // 2

                with PILImage.open(img_path) as img:
                    resized = img.convert("RGB").resize((fit_w_px, fit_h_px), PILImage.LANCZOS)

                tmp_path = str(Path(tmp_dir) / f"_tmp_{Path(img_path).stem}.jpg")
                resized.save(tmp_path, "JPEG", quality=95)

                xl_img = XLImage(tmp_path)

                from_marker = AnchorMarker(
                    col=min_col - 1,
                    colOff=x_off,
                    row=min_row - 1,
                    rowOff=y_off,
                )
                ext = XDRPositiveSize2D(fit_w_emu, fit_h_emu)
                xl_img.anchor = OneCellAnchor(_from=from_marker, ext=ext)
                ws.add_image(xl_img)

                print(f"    スロット{i+1}: section='{slot['section']}' category='{slot['category']}'")
                print(f"      セル(row={min_row},col={min_col}) サイズ=({fit_w_px}x{fit_h_px}px) offset=({x_off},{y_off})")
                print(f"      画像: {Path(img_path).name}")
                total_placed += 1

        wb.save(output_path)
        print(f"\n{'='*60}")
        print(f"完了: {total_placed}枚配置 → {output_path}")
        print(f"{'='*60}")

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)