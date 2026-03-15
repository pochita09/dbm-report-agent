"""
Step 1: テンプレートExcel解析スクリプト
写真スロット検出ロジックの確認用
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import statistics
import json
from pathlib import Path


def get_merged_cell_value(ws, row, col):
    """結合セルを考慮してセルの値を取得"""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return str(cell.value).strip()

    # 結合セル範囲を確認
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            # 結合セルの左上のセルの値を返す
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            if top_left.value is not None:
                return str(top_left.value).strip()

    return ""


def get_row_height(ws, row):
    """行の高さを取得（デフォルト値を使用）"""
    rd = ws.row_dimensions.get(row)
    if rd and rd.height is not None:
        return rd.height
    return 15.0  # Excelのデフォルト行高さ


def get_col_width(ws, col):
    """列の幅を取得（デフォルト値を使用）"""
    col_letter = openpyxl.utils.get_column_letter(col)
    cd = ws.column_dimensions.get(col_letter)
    if cd and cd.width is not None:
        return cd.width
    return 8.43  # Excelのデフォルト列幅


def detect_photo_slots(ws):
    """
    写真スロットを検出する

    ロジック:
    1. 全行の高さを取得し、中央値の3倍以上の行を「写真行」として検出
    2. 全列の幅を取得し、上位N列を「コンテンツ列」として検出
    3. 写真行×コンテンツ列の交差が写真スロット
    4. 各スロットの直上セルからカテゴリ名、さらに上からセクション名を取得
    """
    max_row = ws.max_row
    max_col = ws.max_column

    print(f"\n  シート: '{ws.title}'  (max_row={max_row}, max_col={max_col})")

    # --- 行高さ分析 ---
    row_heights = {}
    for r in range(1, max_row + 1):
        row_heights[r] = get_row_height(ws, r)

    heights = list(row_heights.values())
    if not heights:
        return []

    median_h = statistics.median(heights)
    threshold_h = median_h * 3.0

    photo_rows = [r for r, h in row_heights.items() if h >= threshold_h]

    print(f"  行高さ中央値: {median_h:.1f}pt  閾値: {threshold_h:.1f}pt")
    print(f"  写真行候補: {photo_rows}")

    if not photo_rows:
        print("  → 写真行なし。スキップ")
        return []

    # --- 列幅分析 ---
    col_widths = {}
    for c in range(1, max_col + 1):
        col_widths[c] = get_col_width(ws, c)

    widths = list(col_widths.values())
    median_w = statistics.median(widths)
    threshold_w = median_w * 1.2  # 中央値の1.2倍以上をコンテンツ列とする

    content_cols = [c for c, w in col_widths.items() if w >= threshold_w]

    print(f"  列幅中央値: {median_w:.1f}  閾値: {threshold_w:.1f}")
    print(f"  コンテンツ列候補: {content_cols}")

    if not content_cols:
        print("  → コンテンツ列なし。スキップ")
        return []

    # --- スロット検出 ---
    slots = []

    # 各写真行の探索上限（前の写真行まで）
    sorted_photo_rows = sorted(photo_rows)

    for idx, r in enumerate(sorted_photo_rows):
        # この写真行グループの上限（前の写真行+1、なければ行1）
        row_min = sorted_photo_rows[idx - 1] + 1 if idx > 0 else 1

        for c in content_cols:
            # カテゴリ名：直上セル（写真行-1から上限まで）
            category = ""
            cat_row = r  # カテゴリが見つかった行
            for look_up in range(1, r - row_min + 2):  # 上限まで探索
                search_row = r - look_up
                if search_row < row_min:
                    break
                candidate = get_merged_cell_value(ws, search_row, c)
                if candidate:
                    category = candidate
                    cat_row = search_row
                    break

            # セクション名：カテゴリより上（同じグループ内のみ）
            # col=1 も含めて横断的にセクション名を探す（結合セルで別列に書かれている場合のため）
            section = ""
            for search_row in range(cat_row - 1, row_min - 1, -1):
                # まず同じ列で探す
                candidate = get_merged_cell_value(ws, search_row, c)
                if not candidate:
                    # col=1（最左列）でも探す（セクション見出しが左に寄っている場合）
                    candidate = get_merged_cell_value(ws, search_row, content_cols[0])
                if candidate and candidate != category:
                    section = candidate
                    break

            slot = {
                "row": r,
                "col": c,
                "row_height": row_heights[r],
                "col_width": col_widths[c],
                "category": category,
                "section": section,
            }
            slots.append(slot)
            print(f"  スロット: row={r}, col={c} | セクション='{section}' | カテゴリ='{category}' | 高さ={row_heights[r]:.1f}pt, 幅={col_widths[c]:.1f}")

    return slots


def analyze_workbook(filepath):
    """Excelファイル全体を解析"""
    print(f"\n{'='*60}")
    print(f"ファイル: {filepath}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"シート一覧: {wb.sheetnames}")

    all_slots = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        slots = detect_photo_slots(ws)
        if slots:
            all_slots[sheet_name] = slots
            print(f"  [OK] '{sheet_name}': {len(slots)}スロット検出")
        else:
            print(f"  [--] '{sheet_name}': 写真スロットなし（スキップ）")

    print(f"\n{'='*60}")
    print(f"結果サマリー:")
    print(f"  写真ページ: {list(all_slots.keys())}")
    total = sum(len(s) for s in all_slots.values())
    print(f"  総スロット数: {total}")
    print(f"{'='*60}")

    return all_slots


if __name__ == "__main__":
    template_path = r"c:\Users\nyaaa\OneDrive\デスクトップ\報告書テンプレ\報告書フォーマット.xlsx"

    result = analyze_workbook(template_path)

    print("\n詳細JSON出力:")
    print(json.dumps(
        {sheet: [
            {"row": s["row"], "col": s["col"], "section": s["section"], "category": s["category"]}
            for s in slots
        ] for sheet, slots in result.items()},
        ensure_ascii=False,
        indent=2
    ))
